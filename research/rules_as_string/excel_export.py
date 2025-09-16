"""
Excel Export Module for Sentinel-2 Analysis
Exports all database tables to Excel with multiple worksheets and timestamps
"""
import sqlite3
from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelExporter:
    def __init__(self, db_path='sentinel_analysis.db'):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
    
    def close_connection(self):
        """Close database connection"""
        self.conn.close()
    
    def get_timestamp_filename(self, base_name="sentinel_analysis"):
        """Generate filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_{timestamp}.xlsx"
    
    def get_table_data(self, table_name):
        """Get all data from a specific table"""
        self.cursor.execute(f"SELECT * FROM {table_name}")
        columns = [description[0] for description in self.cursor.description]
        data = self.cursor.fetchall()
        return columns, data
    
    def format_worksheet(self, ws, title):
        """Format worksheet with headers and styling"""
        # Set title
        ws.title = title
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Style for data
        data_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Apply data formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_worksheet(self, wb):
        """Create a summary worksheet with statistics"""
        ws = wb.create_sheet("Summary", 0)  # Insert at beginning
        
        # Get statistics
        self.cursor.execute("SELECT COUNT(*) FROM Sentinel_hub_calls_data")
        sentinel_count = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(*) FROM Thresholds")
        threshold_count = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(*) FROM Transactions")
        transaction_count = self.cursor.fetchone()[0]
        
        # Get unique plots
        self.cursor.execute("SELECT COUNT(DISTINCT plot_no) FROM Sentinel_hub_calls_data")
        unique_plots = self.cursor.fetchone()[0]
        
        # Get date range
        self.cursor.execute("SELECT MIN(date), MAX(date) FROM Sentinel_hub_calls_data")
        date_range = self.cursor.fetchone()
        
        # Create summary data
        summary_data = [
            ["Sentinel-2 Satellite Data Analysis Summary", ""],
            ["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Database Statistics", ""],
            ["Sentinel Data Records:", sentinel_count],
            ["Threshold Rules:", threshold_count],
            ["Processed Transactions:", transaction_count],
            ["Unique Plots:", unique_plots],
            ["Date Range:", f"{date_range[0]} to {date_range[1]}"],
            ["", ""],
            ["Worksheet Contents", ""],
            ["1. Summary", "This overview worksheet"],
            ["2. Sentinel_Data", "Raw satellite data (NDVI, NDBI, NDWI)"],
            ["3. Thresholds", "Land type rules and thresholds"],
            ["4. Transactions", "Analysis results and inferences"],
            ["5. Analysis_Stats", "Statistical analysis summary"]
        ]
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Title
                    cell.font = Font(bold=True, size=16)
                elif row_idx in [4, 11]:  # Section headers
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_analysis_stats_worksheet(self, wb):
        """Create analysis statistics worksheet"""
        ws = wb.create_sheet("Analysis_Stats")
        
        # Get transaction data
        self.cursor.execute("SELECT * FROM Transactions")
        transactions = self.cursor.fetchall()
        
        if not transactions:
            ws.cell(row=1, column=1, value="No transaction data available")
            return
        
        # Calculate statistics
        veg_changes = sum(1 for t in transactions if t[16] != "No Change")
        con_changes = sum(1 for t in transactions if t[17] != "No Change")
        flood_changes = sum(1 for t in transactions if t[18] != "No Change")
        
        # Count by change type
        veg_types = {}
        con_types = {}
        flood_types = {}
        
        for trans in transactions:
            veg = trans[16]
            con = trans[17]
            flood = trans[18]
            
            veg_types[veg] = veg_types.get(veg, 0) + 1
            con_types[con] = con_types.get(con, 0) + 1
            flood_types[flood] = flood_types.get(flood, 0) + 1
        
        # Create statistics data
        stats_data = [
            ["Analysis Statistics", ""],
            ["", ""],
            ["Change Detection Summary", ""],
            ["Total Transactions:", len(transactions)],
            ["Vegetation Changes:", veg_changes],
            ["Construction Changes:", con_changes],
            ["Flooding Changes:", flood_changes],
            ["", ""],
            ["Vegetation Change Types", "Count"],
        ]
        
        # Add vegetation change types
        for change_type, count in sorted(veg_types.items(), key=lambda x: x[1], reverse=True):
            stats_data.append([change_type, count])
        
        stats_data.extend([
            ["", ""],
            ["Construction Change Types", "Count"],
        ])
        
        # Add construction change types
        for change_type, count in sorted(con_types.items(), key=lambda x: x[1], reverse=True):
            stats_data.append([change_type, count])
        
        stats_data.extend([
            ["", ""],
            ["Flooding Change Types", "Count"],
        ])
        
        # Add flooding change types
        for change_type, count in sorted(flood_types.items(), key=lambda x: x[1], reverse=True):
            stats_data.append([change_type, count])
        
        # Write statistics data
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in [1, 3, 9, 15, 21]:  # Headers
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_excel(self, output_file=None):
        """Export all tables to Excel with multiple worksheets"""
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl library not installed. Install it with: pip install openpyxl")
            return None
        
        if output_file is None:
            output_file = self.get_timestamp_filename()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        try:
            # Create summary worksheet
            self.create_summary_worksheet(wb)
            
            # Export Sentinel_hub_calls_data
            columns, data = self.get_table_data("Sentinel_hub_calls_data")
            ws = wb.create_sheet("Sentinel_Data")
            
            # Write headers
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            self.format_worksheet(ws, "Sentinel_Data")
            
            # Export Thresholds
            columns, data = self.get_table_data("Thresholds")
            ws = wb.create_sheet("Thresholds")
            
            # Write headers
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            self.format_worksheet(ws, "Thresholds")
            
            # Export Transactions
            columns, data = self.get_table_data("Transactions")
            ws = wb.create_sheet("Transactions")
            
            # Write headers
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            self.format_worksheet(ws, "Transactions")
            
            # Create analysis statistics worksheet
            self.create_analysis_stats_worksheet(wb)
            
            # Save workbook
            wb.save(output_file)
            print(f"✅ Excel file exported successfully: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            return None
    
    def close_connection(self):
        """Close database connection"""
        self.conn.close()

def main():
    """Main function to demonstrate Excel export"""
    print("📊 Excel Export Demo for Sentinel-2 Analysis")
    print("=" * 50)
    
    # Check if database exists
    if not os.path.exists('sentinel_analysis.db'):
        print("❌ Database not found. Please run database_schema.py first.")
        return
    
    # Create exporter and export
    exporter = ExcelExporter()
    
    try:
        output_file = exporter.export_to_excel()
        if output_file:
            print(f"\n📁 Excel file created: {output_file}")
            
            # Show file info
            if os.path.exists(output_file):
                size = os.path.getsize(output_file)
                print(f"📏 File size: {size:,} bytes")
            
            print("\n📋 Worksheets included:")
            print("1. Summary - Overview and statistics")
            print("2. Sentinel_Data - Raw satellite data")
            print("3. Thresholds - Land type rules")
            print("4. Transactions - Analysis results")
            print("5. Analysis_Stats - Statistical summary")
    
    finally:
        exporter.close_connection()

if __name__ == "__main__":
    main()
